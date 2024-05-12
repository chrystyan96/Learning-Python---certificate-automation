# Lib para carregar arquivos do excel - pip install openpyxl
import openpyxl as xls

# Lib para alterar a img do certificado - pip install pillow
from PIL import Image, ImageDraw, ImageFont

# Abrindo a planilha
workbook_alunos = xls.load_workbook('planilha_alunos.xlsx')
sheet_alunos = workbook_alunos['Sheet1']

# Extaindo informações da planilha
# Começa a ler (iter_rows()) a partir da linha 2 da planilha (min_row=2)
for indice, linha in enumerate(sheet_alunos.iter_rows(min_row=2)):
    # Acessando e salvando o valor de cada célula
    nome_curso = linha[0].value             # Nome do curso 
    nome_participante = linha[1].value      # Nome do participante
    tipo_participacao = linha[2].value      # Tipo de participação
    data_inicio = linha[3].value            # Data de início do curso
    data_termino = linha[4].value           # Data de término do curso
    carga_horaria = linha[5].value          # Carga horária do curso
    data_emissao = linha[6].value           # Data de emissão do certificado

    # Transferir os dados da planilha para a img do certificado
    # Configurando as fontes a serem usadas
    fonte_nome = ImageFont.truetype('./tahomabd.ttf', 90)
    fonte_geral = ImageFont.truetype('./tahoma.ttf', 80)
    fonte_datas = ImageFont.truetype('./tahoma.ttf', 60)
    
    # Acessando a img do certificado
    imagem = Image.open('./certificado_padrao.jpg')
    desenhar = ImageDraw.Draw(imagem)
    
    # Adicionando texto na img passando as coordenadas, o texto, a cor e a fonte como parâmentros
    desenhar.text((1020, 830), nome_participante, fill='black', font=fonte_nome)
    desenhar.text((1080, 950), nome_curso, fill='black', font=fonte_geral)
    desenhar.text((1440, 1070), tipo_participacao, fill='black', font=fonte_geral)
    desenhar.text((1490, 1190), f'{str(carga_horaria)} horas', fill='black', font=fonte_geral)
    desenhar.text((735, 1780), data_inicio, fill='black', font=fonte_datas)
    desenhar.text((735, 1930), data_termino, fill='black', font=fonte_datas)
    desenhar.text((2210, 1930), data_emissao, fill='black', font=fonte_datas)

    
    # Salvando a img alterada
    imagem.save(f'./certificados_automatizados/{nome_participante.replace(" ", "_")}_certificado.png')